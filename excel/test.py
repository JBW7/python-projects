from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# CREATE A WORKBOOK OBJECT
wb = Workbook()

# LOAD EXISTING WORKSHEET
wb = load_workbook('names.xlsx')

# CREATE A ACTIVE WORKSHEET
ws = wb.active

# PRINT FROM SPREADSHEET
print (f'{ws['B3'].value} : {ws['B4'].value')